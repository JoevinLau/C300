from __future__ import annotations

import hashlib
import json
import math
import os
import re
import threading
import uuid
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable


SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}
DEFAULT_EMBEDDING_MODEL = "text-embedding-3-small"
DEFAULT_TOP_K = 6
DEFAULT_SCORE_THRESHOLD = 0.25
INDEX_SCHEMA_VERSION = 1


class RagError(Exception):
    pass


class UnsupportedDocumentError(RagError):
    pass


class EmptyDocumentError(RagError):
    pass


@dataclass(frozen=True)
class ExtractedSection:
    text: str
    location: str
    page: int | None = None
    sheet: str | None = None
    row_start: int | None = None
    row_end: int | None = None


@dataclass(frozen=True)
class DocumentRecord:
    document_id: str
    filename: str
    file_type: str
    content_hash: str
    chunk_count: int
    status: str = "indexed"
    error: str | None = None


@dataclass(frozen=True)
class SearchResult:
    document_id: str
    filename: str
    location: str
    excerpt: str
    score: float


@dataclass(frozen=True)
class VectorRecord:
    id: str
    text: str
    embedding: list[float]
    metadata: dict[str, str | int]


@dataclass(frozen=True)
class WorkspaceIndex:
    schema_version: int
    documents: list[DocumentRecord]
    vectors: list[VectorRecord]


def _safe_workspace_id(workspace_id: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9._-]", "-", workspace_id.strip())
    if not normalized or normalized in {".", ".."}:
        raise RagError("workspace_id must contain at least one letter or number")
    return normalized[:100]


def _content_hash(contents: bytes) -> str:
    return hashlib.sha256(contents).hexdigest()


def _tokenizer() -> Any:
    try:
        import tiktoken

        return tiktoken.get_encoding("cl100k_base")
    except ImportError as exc:
        raise RagError(
            "Python dependency tiktoken is missing. Install api/requirements.txt."
        ) from exc


def chunk_text(
    text: str,
    *,
    target_tokens: int = 600,
    overlap_tokens: int = 100,
) -> list[str]:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if not cleaned:
        return []

    encoding = _tokenizer()
    tokens = encoding.encode(cleaned)
    if len(tokens) <= target_tokens:
        return [cleaned]

    chunks: list[str] = []
    step = max(1, target_tokens - overlap_tokens)
    for start in range(0, len(tokens), step):
        token_slice = tokens[start : start + target_tokens]
        if not token_slice:
            break
        chunk = encoding.decode(token_slice).strip()
        if chunk:
            chunks.append(chunk)
        if start + target_tokens >= len(tokens):
            break
    return chunks


def extract_pdf(contents: bytes) -> list[ExtractedSection]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RagError(
            "Python dependency pypdf is missing. Install api/requirements.txt."
        ) from exc

    try:
        reader = PdfReader(BytesIO(contents))
        sections = []
        for page_number, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                sections.append(
                    ExtractedSection(
                        text=text,
                        location=f"page {page_number}",
                        page=page_number,
                    )
                )
        return sections
    except Exception as exc:
        raise RagError(f"Failed to parse PDF: {exc}") from exc


def extract_xlsx(contents: bytes, *, rows_per_section: int = 25) -> list[ExtractedSection]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RagError(
            "Python dependency openpyxl is missing. Install api/requirements.txt."
        ) from exc

    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        sections: list[ExtractedSection] = []
        for sheet in workbook.worksheets:
            rows = [
                [str(cell) if cell is not None else "" for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:] or [rows[0]]
            for offset in range(0, len(data_rows), rows_per_section):
                batch = data_rows[offset : offset + rows_per_section]
                row_start = offset + 2 if len(rows) > 1 else 1
                row_end = row_start + len(batch) - 1
                lines = [
                    f"Sheet: {sheet.title}",
                    "Columns: " + " | ".join(header),
                    *[" | ".join(row) for row in batch],
                ]
                sections.append(
                    ExtractedSection(
                        text="\n".join(lines),
                        location=f"sheet {sheet.title}, rows {row_start}-{row_end}",
                        sheet=sheet.title,
                        row_start=row_start,
                        row_end=row_end,
                    )
                )
        workbook.close()
        return sections
    except Exception as exc:
        raise RagError(f"Failed to parse spreadsheet: {exc}") from exc


def extract_xls(contents: bytes, *, rows_per_section: int = 25) -> list[ExtractedSection]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RagError(
            "Python dependency pandas is missing. Install api/requirements.txt."
        ) from exc

    try:
        sheets = pd.read_excel(BytesIO(contents), sheet_name=None, header=None)
    except ImportError as exc:
        raise RagError(
            "Python dependency xlrd is missing. Install api/requirements.txt."
        ) from exc
    except Exception as exc:
        raise RagError(f"Failed to parse spreadsheet: {exc}") from exc

    sections: list[ExtractedSection] = []
    for sheet_name, frame in sheets.items():
        frame = frame.fillna("")
        rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in frame.to_numpy().tolist()
        ]
        rows = [row for row in rows if any(cell.strip() for cell in row)]
        if not rows:
            continue

        header = rows[0]
        data_rows = rows[1:] or [rows[0]]
        for offset in range(0, len(data_rows), rows_per_section):
            batch = data_rows[offset : offset + rows_per_section]
            row_start = offset + 2 if len(rows) > 1 else 1
            row_end = row_start + len(batch) - 1
            lines = [
                f"Sheet: {sheet_name}",
                "Columns: " + " | ".join(header),
                *[" | ".join(row) for row in batch],
            ]
            sections.append(
                ExtractedSection(
                    text="\n".join(lines),
                    location=f"sheet {sheet_name}, rows {row_start}-{row_end}",
                    sheet=str(sheet_name),
                    row_start=row_start,
                    row_end=row_end,
                )
            )
    return sections


def extract_document(filename: str, contents: bytes) -> list[ExtractedSection]:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise UnsupportedDocumentError(
            f"Unsupported file type '{extension or 'unknown'}'. Upload PDF, XLSX, or XLS files."
        )
    if not contents:
        raise EmptyDocumentError("The uploaded file is empty.")

    if extension == ".pdf":
        sections = extract_pdf(contents)
    elif extension == ".xls":
        sections = extract_xls(contents)
    else:
        sections = extract_xlsx(contents)
    if not sections:
        raise EmptyDocumentError("No readable text or spreadsheet rows were found.")
    return sections


def _chunk_sections(sections: Iterable[ExtractedSection]) -> list[tuple[str, ExtractedSection]]:
    chunks: list[tuple[str, ExtractedSection]] = []
    for section in sections:
        for text in chunk_text(section.text):
            chunks.append((text, section))
    return chunks


class RagService:
    def __init__(
        self,
        data_root: Path | str | None = None,
        *,
        openai_client: Any | None = None,
    ) -> None:
        configured_root = data_root or os.getenv("RAG_DATA_DIR")
        self.data_root = Path(configured_root or Path.home() / ".c300" / "rag")
        self.embedding_model = os.getenv(
            "RAG_EMBEDDING_MODEL", DEFAULT_EMBEDDING_MODEL
        )
        self.top_k = int(os.getenv("RAG_TOP_K", str(DEFAULT_TOP_K)))
        self.score_threshold = float(
            os.getenv("RAG_SCORE_THRESHOLD", str(DEFAULT_SCORE_THRESHOLD))
        )
        self._openai_client = openai_client
        self._lock = threading.RLock()

    def _workspace_dir(self, workspace_id: str) -> Path:
        workspace_dir = self.data_root / "workspaces" / _safe_workspace_id(workspace_id)
        workspace_dir.mkdir(parents=True, exist_ok=True)
        return workspace_dir

    def _legacy_manifest_path(self, workspace_id: str) -> Path:
        return self._workspace_dir(workspace_id) / "documents.json"

    def _legacy_vectors_path(self, workspace_id: str) -> Path:
        return self._workspace_dir(workspace_id) / "vectors.json"

    def _index_path(self, workspace_id: str) -> Path:
        return self._workspace_dir(workspace_id) / "index.json"

    def _backup_index_path(self, workspace_id: str) -> Path:
        return self._workspace_dir(workspace_id) / "index.backup.json"

    def _load_legacy_manifest(self, workspace_id: str) -> list[DocumentRecord]:
        path = self._legacy_manifest_path(workspace_id)
        if not path.exists():
            return []
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            return [DocumentRecord(**item) for item in payload]
        except (OSError, ValueError, TypeError) as exc:
            raise RagError(f"Failed to read legacy document index: {exc}") from exc

    def _load_legacy_vectors(self, workspace_id: str) -> list[VectorRecord]:
        path = self._legacy_vectors_path(workspace_id)
        if not path.exists():
            return []
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            return [VectorRecord(**item) for item in payload]
        except (OSError, ValueError, TypeError) as exc:
            raise RagError(f"Failed to read legacy vector index: {exc}") from exc

    def _validate_index(self, index: WorkspaceIndex) -> None:
        if index.schema_version != INDEX_SCHEMA_VERSION:
            raise RagError(
                "Unsupported RAG index schema version "
                f"{index.schema_version}; expected {INDEX_SCHEMA_VERSION}."
            )

        documents_by_id: dict[str, DocumentRecord] = {}
        for document in index.documents:
            if document.document_id in documents_by_id:
                raise RagError(
                    f"RAG index contains duplicate document id {document.document_id}."
                )
            if document.chunk_count < 1:
                raise RagError(
                    f"RAG document {document.document_id} has no indexed chunks."
                )
            documents_by_id[document.document_id] = document

        vector_ids: set[str] = set()
        chunk_counts = {document_id: 0 for document_id in documents_by_id}
        for vector in index.vectors:
            if vector.id in vector_ids:
                raise RagError(f"RAG index contains duplicate vector id {vector.id}.")
            vector_ids.add(vector.id)

            document_id = str(vector.metadata.get("document_id", ""))
            document = documents_by_id.get(document_id)
            if document is None:
                raise RagError(
                    f"RAG vector {vector.id} references missing document {document_id}."
                )
            if str(vector.metadata.get("filename", "")) != document.filename:
                raise RagError(
                    f"RAG vector {vector.id} filename does not match its document."
                )
            if not str(vector.metadata.get("location", "")).strip():
                raise RagError(f"RAG vector {vector.id} has no source location.")
            chunk_counts[document_id] += 1

        for document_id, document in documents_by_id.items():
            if chunk_counts[document_id] != document.chunk_count:
                raise RagError(
                    f"RAG document {document_id} expects {document.chunk_count} chunks "
                    f"but has {chunk_counts[document_id]}."
                )

    def _decode_index(self, path: Path) -> WorkspaceIndex:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(payload, dict):
                raise TypeError("root must be an object")
            documents = [
                DocumentRecord(**item) for item in payload.get("documents", [])
            ]
            vectors = [VectorRecord(**item) for item in payload.get("vectors", [])]
            index = WorkspaceIndex(
                schema_version=payload.get("schema_version"),
                documents=documents,
                vectors=vectors,
            )
            self._validate_index(index)
            return index
        except (AttributeError, OSError, ValueError, TypeError) as exc:
            raise RagError(f"Failed to read RAG index {path.name}: {exc}") from exc

    def _reconcile_legacy_index(self, workspace_id: str) -> WorkspaceIndex:
        documents = self._load_legacy_manifest(workspace_id)
        vectors = self._load_legacy_vectors(workspace_id)
        vectors_by_document: dict[str, list[VectorRecord]] = {}
        for vector in vectors:
            document_id = str(vector.metadata.get("document_id", ""))
            vectors_by_document.setdefault(document_id, []).append(vector)

        reconciled_documents = []
        complete_document_ids = set()
        for document in documents:
            document_vectors = vectors_by_document.get(document.document_id, [])
            if len(document_vectors) == document.chunk_count:
                reconciled_documents.append(document)
                complete_document_ids.add(document.document_id)

        reconciled_vectors = [
            vector
            for vector in vectors
            if str(vector.metadata.get("document_id", "")) in complete_document_ids
        ]
        index = WorkspaceIndex(
            schema_version=INDEX_SCHEMA_VERSION,
            documents=reconciled_documents,
            vectors=reconciled_vectors,
        )
        self._validate_index(index)
        return index

    def _load_index(self, workspace_id: str) -> WorkspaceIndex:
        primary_path = self._index_path(workspace_id)
        if primary_path.exists():
            try:
                return self._decode_index(primary_path)
            except RagError as primary_error:
                backup_path = self._backup_index_path(workspace_id)
                if backup_path.exists():
                    try:
                        return self._decode_index(backup_path)
                    except RagError as backup_error:
                        raise RagError(
                            f"Primary and backup RAG indexes are invalid: "
                            f"{primary_error} {backup_error}"
                        ) from backup_error
                raise primary_error

        legacy_manifest = self._legacy_manifest_path(workspace_id)
        legacy_vectors = self._legacy_vectors_path(workspace_id)
        if legacy_manifest.exists() or legacy_vectors.exists():
            index = self._reconcile_legacy_index(workspace_id)
            self._save_index(workspace_id, index.documents, index.vectors)
            return index

        return WorkspaceIndex(INDEX_SCHEMA_VERSION, [], [])

    @staticmethod
    def _serialize_index(index: WorkspaceIndex) -> str:
        return json.dumps(
            {
                "schema_version": index.schema_version,
                "documents": [asdict(document) for document in index.documents],
                "vectors": [asdict(vector) for vector in index.vectors],
            },
            indent=2,
        )

    @staticmethod
    def _atomic_write(path: Path, contents: str) -> None:
        temporary_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
        try:
            with temporary_path.open("w", encoding="utf-8") as handle:
                handle.write(contents)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temporary_path, path)
            if os.name != "nt":
                directory_fd = os.open(path.parent, os.O_RDONLY)
                try:
                    os.fsync(directory_fd)
                finally:
                    os.close(directory_fd)
        finally:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass

    def _save_index(
        self,
        workspace_id: str,
        documents: list[DocumentRecord],
        vectors: list[VectorRecord],
    ) -> None:
        index = WorkspaceIndex(INDEX_SCHEMA_VERSION, documents, vectors)
        self._validate_index(index)
        serialized = self._serialize_index(index)
        primary_path = self._index_path(workspace_id)
        backup_path = self._backup_index_path(workspace_id)

        try:
            if primary_path.exists():
                try:
                    self._decode_index(primary_path)
                except RagError:
                    pass
                else:
                    self._atomic_write(
                        backup_path, primary_path.read_text(encoding="utf-8")
                    )
            else:
                self._atomic_write(backup_path, serialized)
            self._atomic_write(primary_path, serialized)
        except OSError as exc:
            raise RagError(f"Failed to commit RAG workspace index: {exc}") from exc

    def close(self) -> None:
        return None

    def _openai(self) -> Any:
        if self._openai_client is not None:
            return self._openai_client
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise RagError(
                "Python dependency openai is missing. Install api/requirements.txt."
            ) from exc
        self._openai_client = OpenAI(api_key=_get_ai_key())
        return self._openai_client

    def _embed(self, texts: list[str]) -> list[list[float]]:
        if not texts:
            return []
        try:
            response = self._openai().embeddings.create(
                model=self.embedding_model,
                input=texts,
                encoding_format="float",
            )
            ordered = sorted(response.data, key=lambda item: item.index)
            return [item.embedding for item in ordered]
        except Exception as exc:
            raise RagError(f"Embedding request failed: {exc}") from exc

    def list_documents(self, workspace_id: str) -> list[DocumentRecord]:
        with self._lock:
            return list(self._load_index(workspace_id).documents)

    def ingest(
        self, workspace_id: str, filename: str, contents: bytes
    ) -> DocumentRecord:
        content_hash = _content_hash(contents)
        extension = Path(filename).suffix.lower()

        with self._lock:
            index = self._load_index(workspace_id)
            documents = index.documents
            duplicate = next(
                (
                    document
                    for document in documents
                    if document.filename == filename
                    and document.content_hash == content_hash
                ),
                None,
            )
            if duplicate:
                return duplicate

            sections = extract_document(filename, contents)
            chunks = _chunk_sections(sections)
            if not chunks:
                raise EmptyDocumentError("No indexable content was found.")

            embeddings = self._embed([text for text, _section in chunks])
            if len(embeddings) != len(chunks):
                raise RagError("Embedding response count did not match the document chunks.")

            replaced = [
                document for document in documents if document.filename == filename
            ]
            replaced_ids = {document.document_id for document in replaced}
            vectors = [
                vector
                for vector in index.vectors
                if str(vector.metadata.get("document_id")) not in replaced_ids
            ]

            document_id = str(uuid.uuid4())
            ids = [f"{document_id}:{index}" for index in range(len(chunks))]
            metadatas = []
            for index, (_text, section) in enumerate(chunks):
                metadata: dict[str, str | int] = {
                    "document_id": document_id,
                    "filename": filename,
                    "file_type": extension.lstrip("."),
                    "location": section.location,
                    "chunk_index": index,
                    "content_hash": content_hash,
                }
                if section.page is not None:
                    metadata["page"] = section.page
                if section.sheet is not None:
                    metadata["sheet"] = section.sheet
                if section.row_start is not None:
                    metadata["row_start"] = section.row_start
                if section.row_end is not None:
                    metadata["row_end"] = section.row_end
                metadatas.append(metadata)

            vectors.extend(
                VectorRecord(
                    id=ids[index],
                    text=text,
                    embedding=embeddings[index],
                    metadata=metadatas[index],
                )
                for index, (text, _section) in enumerate(chunks)
            )
            record = DocumentRecord(
                document_id=document_id,
                filename=filename,
                file_type=extension.lstrip("."),
                content_hash=content_hash,
                chunk_count=len(chunks),
            )
            remaining = [
                document for document in documents if document.filename != filename
            ]
            self._save_index(workspace_id, [*remaining, record], vectors)
            return record

    def delete_document(self, workspace_id: str, document_id: str) -> bool:
        with self._lock:
            index = self._load_index(workspace_id)
            documents = index.documents
            if not any(
                document.document_id == document_id for document in documents
            ):
                return False
            self._save_index(
                workspace_id,
                [
                    document
                    for document in documents
                    if document.document_id != document_id
                ],
                [
                    vector
                    for vector in index.vectors
                    if str(vector.metadata.get("document_id")) != document_id
                ],
            )
            return True

    def search(self, workspace_id: str, query: str) -> list[SearchResult]:
        if not query.strip():
            return []

        with self._lock:
            index = self._load_index(workspace_id)
        if not index.documents:
            return []

        query_embedding = self._embed([query])[0]
        scored = sorted(
            (
                (_cosine_similarity(query_embedding, vector.embedding), vector)
                for vector in index.vectors
            ),
            key=lambda item: item[0],
            reverse=True,
        )[: self.top_k]
        matches = []
        for score, vector in scored:
            if score < self.score_threshold:
                continue
            metadata = vector.metadata
            matches.append(
                SearchResult(
                    document_id=str(metadata["document_id"]),
                    filename=str(metadata["filename"]),
                    location=str(metadata["location"]),
                    excerpt=vector.text[:500],
                    score=round(score, 4),
                )
            )
        return matches


def _cosine_similarity(left: list[float], right: list[float]) -> float:
    if not left or not right or len(left) != len(right):
        return 0.0
    dot = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(a * a for a in left))
    right_norm = math.sqrt(sum(b * b for b in right))
    if left_norm == 0 or right_norm == 0:
        return 0.0
    return max(0.0, min(1.0, dot / (left_norm * right_norm)))


def _get_ai_key() -> str:
    for env_name in ("AI_KEY", "OPENAI_API_KEY"):
        value = os.getenv(env_name)
        if value and value.strip():
            return value.strip()
    raise RagError("AI_KEY or OPENAI_API_KEY environment variable is not set.")
