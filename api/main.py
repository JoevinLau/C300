# main.py
import os
import sys
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field, model_validator
from ecotransit_scraper import calculate_ecotransit

API_DIR = Path(__file__).resolve().parent
ROOT_DIR = API_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from calculation.method2_calculations import compute_method2, list_machine_library

from service import (
    list_naics_options, 
    compute_emissions, 
    fetch_naics_for_material, 
    save_material_mapping,
)
from dotenv import load_dotenv

load_dotenv(ROOT_DIR / ".env")
load_dotenv(API_DIR / ".env", override=True)

AI_KEY_ENV_NAMES = ("AI_KEY", "OPENAI_API_KEY")

def get_ai_key() -> str:
    for env_name in AI_KEY_ENV_NAMES:
        key = os.environ.get(env_name)
        if key and key.strip():
            return key.strip()

    accepted_names = " or ".join(AI_KEY_ENV_NAMES)
    raise HTTPException(
        status_code=500,
        detail=(
            f"{accepted_names} environment variable not set. "
            "Add AI_KEY=your_key_here or OPENAI_API_KEY=your_key_here "
            "to .env in the project root or api/.env, then restart the API server."
        ),
    )

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(HTTPException)
async def http_exception_handler(_request: Request, exc: HTTPException) -> JSONResponse:
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


@app.exception_handler(Exception)
async def unhandled_exception_handler(_request: Request, exc: Exception) -> JSONResponse:
    return JSONResponse(
        status_code=500,
        content={"detail": f"Internal server error: {exc}"},
    )


class Allocation(BaseModel):
    raw_material_pct: float = Field(..., ge=0)
    fabrication_pct: float = Field(..., ge=0)
    surface_treatment_pct: float = Field(..., ge=0)

    @model_validator(mode="after")
    def validate_allocation_total(self) -> "Allocation":
        total = (
            self.raw_material_pct
            + self.fabrication_pct
            + self.surface_treatment_pct
        )
        if abs(total - 100) > 0.01:
            raise ValueError("Allocation percentages must add up to 100.")
        return self


class Naics(BaseModel):
    raw_material: str = Field(..., min_length=6, max_length=6)
    fabrication: str = Field(..., min_length=6, max_length=6)
    surface_treatment: str = Field(..., min_length=6, max_length=6)


class SgdAmountsInput(BaseModel):
    raw_material: float = Field(..., ge=0)
    fabrication: float = Field(..., ge=0)
    surface_treatment: float = Field(..., ge=0)


class InputData(BaseModel):
    invoice_id: str = Field(..., min_length=1)
    year: int = Field(..., ge=2020, le=2030)
    total_amount_sgd: float = Field(..., gt=0)
    sgd_amounts: SgdAmountsInput | None = None
    allocation: Allocation | None = None
    naics: Naics

    @model_validator(mode="before")
    @classmethod
    def normalize_amounts(cls, data: object) -> object:
        if not isinstance(data, dict):
            return data

        total = data.get("total_amount_sgd")
        sgd_amounts = data.get("sgd_amounts")
        allocation = data.get("allocation")

        if sgd_amounts is None and allocation is not None and total is not None:
            data = {**data, "sgd_amounts": {
                "raw_material": total * allocation["raw_material_pct"] / 100.0,
                "fabrication": total * allocation["fabrication_pct"] / 100.0,
                "surface_treatment": total * allocation["surface_treatment_pct"] / 100.0,
            }}

        return data

    @model_validator(mode="after")
    def validate_sgd_amounts_sum(self) -> "InputData":
        if self.sgd_amounts is None:
            raise ValueError("sgd_amounts or allocation is required")

        component_sum = (
            self.sgd_amounts.raw_material
            + self.sgd_amounts.fabrication
            + self.sgd_amounts.surface_treatment
        )
        if abs(component_sum - self.total_amount_sgd) > 0.01:
            raise ValueError(
                "sgd_amounts must sum to total_amount_sgd "
                f"(got {component_sum:.2f} vs {self.total_amount_sgd:.2f})"
            )
        return self


class SgdAmounts(BaseModel):
    raw_material: float
    fabrication: float
    surface_treatment: float


class UsdAmounts(BaseModel):
    raw_material: float
    fabrication: float
    surface_treatment: float


class Usd2022Amounts(BaseModel):
    raw_material: float
    fabrication: float
    surface_treatment: float


class EmissionFactors(BaseModel):
    raw_material: float
    fabrication: float
    surface_treatment: float


class CalculationDetails(BaseModel):
    fx_rate: float
    inflation_index: float
    year: int
    sgd_amounts: SgdAmounts
    usd_amounts: UsdAmounts
    usd2022_amounts: Usd2022Amounts
    factors: EmissionFactors


class CostBreakdown(BaseModel):
    raw_material_usd2022: float
    fabrication_usd2022: float
    surface_treatment_usd2022: float


class EmissionBreakdown(BaseModel):
    raw_material: float
    fabrication: float
    surface_treatment: float
    total: float


class OutputData(BaseModel):
    invoice_id: str
    calculation: CalculationDetails
    costs: CostBreakdown
    emissions: EmissionBreakdown


class NaicsOption(BaseModel):
    code: str
    description: str
    category: str | None = None
    kgco2e_per_usd: float | None = None

class MappingLearnRequest(BaseModel):
    keyword: str
    naics_code: str
    description: str
    category: str


class EcoTransitRequest(BaseModel):
    port_of_loading: str = Field(..., min_length=1)
    port_of_discharge: str = Field("Singapore", min_length=1)
    weight_kg: float = Field(..., gt=0)
    transport_mode: str = Field("sea", pattern="^(sea|land|air|rail|truck|vessel|ship)$")
    origin_country: str | None = None


class EcoTransitTransport(BaseModel):
    origin: str
    port_of_loading: str
    port_of_discharge: str
    weight_kg: float
    chosen_mode: str
    chosen_emissions_kg: float | None = None
    distance_km: float | None = None
    energy_mj: float | None = None
    source: str
    raw: dict


class EcoTransitResponse(BaseModel):
    transport: EcoTransitTransport


class Method2Naics(BaseModel):
    raw_material: str = Field(..., min_length=6, max_length=6)
    surface_treatment: str = Field(..., min_length=6, max_length=6)
    fabrication: str = Field("333517", min_length=6, max_length=6)


class Method2MachiningEntry(BaseModel):
    machine_type: str = Field(..., min_length=1)
    duty_level: str = Field(..., min_length=1)
    operating_hours: float = Field(..., ge=0)


class Method2InputData(BaseModel):
    part_id: str = Field(..., min_length=1)
    year: int = Field(..., ge=2020, le=2030)
    raw_material_sgd: float = Field(..., ge=0)
    surface_treatment_sgd: float = Field(..., ge=0)
    naics: Method2Naics
    transport_emissions_kg: float = Field(0, ge=0)
    transport_source: str = "EcoTransit World"
    machining_entries: list[Method2MachiningEntry] = Field(default_factory=list)

# ---------- ENDPOINTS ----------

@app.get("/fetch-naics")
def get_naics_by_material(name: str):
    return fetch_naics_for_material(name)

@app.post("/learn-mapping")
def learn_mapping(data: MappingLearnRequest):
    save_material_mapping(data.keyword, data.naics_code, data.description, data.category)
    return {"status": "success"}

@app.get("/naics", response_model=list[NaicsOption])
def get_naics_options():
    return list_naics_options()


@app.post("/calculate", response_model=OutputData)
def calculate_emissions(data: InputData):
    payload = {
        "invoice_id": data.invoice_id,
        "year": data.year,
        "total_amount_sgd": data.total_amount_sgd,
        "sgd_amounts": data.sgd_amounts.model_dump(),
        "naics": data.naics.model_dump(),
    }

    try:
        result = compute_emissions(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return OutputData(
        invoice_id=data.invoice_id,
        calculation=CalculationDetails(**result["calculation"]),
        costs=CostBreakdown(**result["costs"]),
        emissions=EmissionBreakdown(**result["emissions"]),
    )


class ChatResponse(BaseModel):
    reply: str


@app.post("/method2-chat", response_model=ChatResponse)
def method2_chat(
    message: str = Form(...),
    excel_file: UploadFile | None = File(None),
):
    key = get_ai_key()

    try:
        from openai import OpenAI
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"OpenAI dependency missing: {exc}. Install requirements with `pip install -r requirements.txt`.",
        ) from exc

    try:
        client = OpenAI(api_key=key)
        content = message
        file_description = None

        if excel_file is not None:
            contents = excel_file.file.read()
            if contents:
                file_description = f"Uploaded spreadsheet filename={excel_file.filename} size={len(contents)} bytes"
                content = f"{file_description}\n\n{message}"
            excel_file.file.close()

        resp = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": content}],
            max_tokens=500,
        )
        choice = resp.choices[0] if resp.choices else None
        text = None
        if choice and hasattr(choice, 'message') and choice.message is not None:
            text = choice.message.get('content') if isinstance(choice.message, dict) else getattr(choice.message, 'content', None)
        if not isinstance(text, str):
            raise ValueError("OpenAI API response did not contain a valid text reply.")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return ChatResponse(reply=text)


@app.post("/method2-chat-file", response_model=ChatResponse)
def method2_chat_file(
    message: str = Form(...),
    excel_file: UploadFile | None = File(None),
):
    key = get_ai_key()

    file_content_description = ""
    if excel_file is not None:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail=f"Python dependency openpyxl is missing: {exc}. Install requirements with `pip install -r requirements.txt`.",
            ) from exc

        try:
            workbook = load_workbook(filename=excel_file.file, data_only=True)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            excel_file.file.close()
            file_content_description = f"Extracted spreadsheet table with {len(rows)} rows and {len(rows[0]) if rows else 0} columns."
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse uploaded Excel file: {exc}") from exc

    try:
        from openai import OpenAI
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"OpenAI dependency missing: {exc}. Install requirements with `pip install -r requirements.txt`.",
        ) from exc

    try:
        client = OpenAI(api_key=key)
        content = message
        if file_content_description:
            content = f"{file_content_description}\n\n{message}"

        resp = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": content}],
            max_tokens=500,
        )
        choice = resp.choices[0] if resp.choices else None
        text = None
        if choice and hasattr(choice, 'message') and choice.message is not None:
            text = choice.message.get('content') if isinstance(choice.message, dict) else getattr(choice.message, 'content', None)
        if not isinstance(text, str):
            raise ValueError("OpenAI API response did not contain a valid text reply.")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return ChatResponse(reply=text)

@app.get("/")
def home():
    return {"message": "API is running!"}


@app.post("/ecotransit")
def ecotransit(data: EcoTransitRequest):
    try:
        result = calculate_ecotransit(
            port_of_loading=data.port_of_loading,
            port_of_discharge=data.port_of_discharge,
            weight_kg=data.weight_kg,
            transport_mode=data.transport_mode,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"EcoTransit scraper failed: {exc}") from exc

    return {
        "transport": {
            "origin": data.origin_country or data.port_of_loading,
            "port_of_loading": data.port_of_loading,
            "port_of_discharge": data.port_of_discharge,
            "weight_kg": data.weight_kg,
            "chosen_mode": data.transport_mode,
            "chosen_emissions_kg": result.get("co2e_kg"),
            "distance_km": result.get("distance_km"),
            "energy_mj": result.get("energy_mj"),
            "source": "EcoTransit World",
            "raw": result,
        }
    }


@app.get("/method2/machines")
def method2_machines():
    return {"machines": list_machine_library()}


@app.post("/method2/calculate")
def calculate_method2(data: Method2InputData):
    payload = {
        "part_id": data.part_id,
        "year": data.year,
        "raw_material_sgd": data.raw_material_sgd,
        "surface_treatment_sgd": data.surface_treatment_sgd,
        "naics": data.naics.model_dump(),
        "transport_emissions_kg": data.transport_emissions_kg,
        "transport_source": data.transport_source,
        "machining_entries": [
            {
                "machine_type": item.machine_type,
                "duty_level": item.duty_level,
                "operating_hours": item.operating_hours,
            }
            for item in data.machining_entries
        ],
    }

    try:
        return compute_method2(payload, spend_calculator=compute_emissions)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)

