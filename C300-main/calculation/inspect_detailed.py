import pandas as pd
import openpyxl

# Check raw_material.xlsx
print("=" * 60)
print("raw_material.xlsx - First 20 rows (no header parsing)")
print("=" * 60)
wb = openpyxl.load_workbook("raw_material.xlsx")
ws = wb.active
print(f"Sheet name: {ws.title}")
print("\nFirst 20 rows of raw data:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    print(f"Row {i}: {row}")

print("\n" + "=" * 60)
print("treatment_cost.xlsx - First 20 rows (no header parsing)")
print("=" * 60)
wb = openpyxl.load_workbook("treatment_cost.xlsx")
ws = wb.active
print(f"Sheet name: {ws.title}")
print("\nFirst 20 rows of raw data:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    print(f"Row {i}: {row}")
