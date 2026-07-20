from __future__ import annotations

import argparse
import hashlib
import math
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
API_DIR = ROOT_DIR / "api"
if str(API_DIR) not in sys.path:
    sys.path.insert(0, str(API_DIR))

DEFAULT_WORKBOOK = ROOT_DIR / "DB" / "Open CEDA 2025.xlsx"
REQUIRED_SHEETS = {
    "Cover",
    "GHG_t_Raw",
    "Exchange rates",
    "Purchaser - producer conversion",
    "Sector level Price Index",
    "Metadata",
}


@dataclass(frozen=True)
class DatasetMetadata:
    version_code: str
    release_date: date | None
    raw_factor_year: int
    reference_price_year: int
    currency_code: str
    price_basis: str
    source_file_sha256: str
    source_license: str
    attribution: str


@dataclass(frozen=True)
class FactorRow:
    country_code: str
    country_name: str
    sector_code: str
    sector_name: str
    sector_description: str | None
    factor_value: float
    raw_factor_value: float
    purchaser_conversion: float
    sector_price_index: float
    exchange_rate_lcu_per_usd: float


def _code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _positive_float(value: Any, label: str) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} is not numeric: {value!r}") from exc
    if not math.isfinite(parsed) or parsed <= 0:
        raise ValueError(f"{label} must be greater than zero: {value!r}")
    return parsed


def _nonnegative_float(value: Any, label: str) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} is not numeric: {value!r}") from exc
    if not math.isfinite(parsed) or parsed < 0:
        raise ValueError(f"{label} must be zero or greater: {value!r}")
    return parsed


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _release_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return datetime.fromisoformat(str(value)).date()
        except ValueError:
            pass
    return None


def _row_map(sheet, key_row: int, value_row: int) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for column in range(1, sheet.max_column + 1):
        key = _code(sheet.cell(key_row, column).value)
        if key:
            result[key] = sheet.cell(value_row, column).value
    return result


def _sector_descriptions(sheet) -> dict[str, tuple[str, str | None]]:
    result: dict[str, tuple[str, str | None]] = {}
    for row in range(1, sheet.max_row + 1):
        code = _code(sheet.cell(row, 1).value)
        name = str(sheet.cell(row, 2).value or "").strip()
        if not code or not name:
            continue
        description = str(sheet.cell(row, 3).value or "").strip() or None
        result[code] = (name, description)
    return result


def extract_open_ceda(
    workbook_path: Path,
    *,
    reference_price_year: int = 2025,
    currency_code: str = "SGD",
    price_basis: str = "purchaser_price",
) -> tuple[DatasetMetadata, list[FactorRow]]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Open CEDA workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Open CEDA workbook is missing sheet(s): {', '.join(missing)}")

    cover = workbook["Cover"]
    raw = workbook["GHG_t_Raw"]
    conversions = workbook["Purchaser - producer conversion"]
    sector_prices = workbook["Sector level Price Index"]
    exchange_rates = workbook["Exchange rates"]
    descriptions = _sector_descriptions(workbook["Metadata"])

    version_code = str(cover["D10"].value or "").strip()
    if not re.fullmatch(r"CEDA\s+\d{4}", version_code, flags=re.IGNORECASE):
        raise ValueError(f"Unexpected Open CEDA version in Cover!D10: {version_code!r}")

    raw_factor_year = int(raw["B1"].value)
    conversion_by_code = _row_map(conversions, 5, 6)

    price_year_row = None
    raw_year_row = None
    for row in range(1, sector_prices.max_row + 1):
        year = sector_prices.cell(row, 1).value
        if _code(year) == str(reference_price_year):
            price_year_row = row
        if _code(year) == str(raw_factor_year):
            raw_year_row = row
    if price_year_row is None or raw_year_row is None:
        raise ValueError(
            f"Sector level Price Index must include {raw_factor_year} and {reference_price_year}."
        )
    reference_price_by_code = _row_map(sector_prices, 5, price_year_row)
    raw_price_by_code = _row_map(sector_prices, 5, raw_year_row)

    exchange_year_column = None
    for column in range(1, exchange_rates.max_column + 1):
        if _code(exchange_rates.cell(4, column).value) == str(reference_price_year):
            exchange_year_column = column
            break
    if exchange_year_column is None:
        raise ValueError(f"Exchange rates does not include {reference_price_year}.")

    singapore_row = None
    for row in range(5, exchange_rates.max_row + 1):
        if _code(exchange_rates.cell(row, 1).value) == "SGP":
            singapore_row = row
            break
    if currency_code != "SGD" or singapore_row is None:
        raise ValueError("Version 1 supports only the Singapore Dollar (SGD).")
    exchange_rate = _positive_float(
        exchange_rates.cell(singapore_row, exchange_year_column).value,
        f"SGD exchange rate for {reference_price_year}",
    )

    sector_columns: list[tuple[int, str, str]] = []
    for column in range(4, raw.max_column + 1):
        sector_code = _code(raw.cell(4, column).value)
        sector_name = str(raw.cell(3, column).value or "").strip()
        if sector_code and sector_name:
            sector_columns.append((column, sector_code, sector_name))
    if not sector_columns:
        raise ValueError("GHG_t_Raw does not contain sector columns in rows 3 and 4.")

    rows: list[FactorRow] = []
    missing_sector_inputs: set[str] = set()
    for row in range(5, raw.max_row + 1):
        country_code = _code(raw.cell(row, 1).value)
        country_name = str(raw.cell(row, 2).value or "").strip()
        if not re.fullmatch(r"[A-Z]{3}", country_code) or not country_name:
            continue

        for column, sector_code, raw_sector_name in sector_columns:
            raw_value = raw.cell(row, column).value
            if raw_value is None:
                continue
            try:
                conversion = _positive_float(
                    conversion_by_code.get(sector_code),
                    f"Purchaser conversion for {sector_code}",
                )
                raw_sector_price = _positive_float(
                    raw_price_by_code.get(sector_code),
                    f"{raw_factor_year} sector price index for {sector_code}",
                )
                reference_sector_price = _positive_float(
                    reference_price_by_code.get(sector_code),
                    f"{reference_price_year} sector price index for {sector_code}",
                )
            except ValueError:
                missing_sector_inputs.add(sector_code)
                continue

            raw_factor = _nonnegative_float(
                raw_value,
                f"Raw factor for {country_code}/{sector_code}",
            )
            factor = (
                raw_factor
                * conversion
                * raw_sector_price
                / reference_sector_price
                / exchange_rate
            )
            metadata_name, description = descriptions.get(
                sector_code,
                (raw_sector_name, None),
            )
            rows.append(
                FactorRow(
                    country_code=country_code,
                    country_name=country_name,
                    sector_code=sector_code,
                    sector_name=metadata_name,
                    sector_description=description,
                    factor_value=factor,
                    raw_factor_value=raw_factor,
                    purchaser_conversion=conversion,
                    sector_price_index=reference_sector_price,
                    exchange_rate_lcu_per_usd=exchange_rate,
                )
            )

    if missing_sector_inputs:
        sample = ", ".join(sorted(missing_sector_inputs)[:10])
        raise ValueError(
            f"{len(missing_sector_inputs)} sector(s) are missing conversion or price-index data: {sample}"
        )
    if not rows:
        raise ValueError("No country-specific Open CEDA emission factors were extracted.")

    metadata = DatasetMetadata(
        version_code=version_code,
        release_date=_release_date(cover["D9"].value),
        raw_factor_year=raw_factor_year,
        reference_price_year=reference_price_year,
        currency_code=currency_code,
        price_basis=price_basis,
        source_file_sha256=_sha256(workbook_path),
        source_license=str(cover["D11"].value or "CC BY-SA 4.0").strip(),
        attribution="CEDA by Watershed",
    )
    return metadata, rows


def _batches(rows: list[Any], size: int) -> Iterable[list[Any]]:
    for start in range(0, len(rows), size):
        yield rows[start : start + size]


def import_open_ceda(metadata: DatasetMetadata, factors: list[FactorRow], batch_size: int) -> int:
    import mysql.connector
    from db import get_conn

    countries = sorted({(row.country_code, row.country_name) for row in factors})
    sectors = sorted(
        {
            (row.sector_code, row.sector_name, row.sector_code, row.sector_description)
            for row in factors
        }
    )
    conn = get_conn()
    try:
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO ceda_dataset_versions
                    (version_code, release_date, raw_factor_year, reference_price_year,
                     currency_code, price_basis, source_name, source_file_sha256,
                     source_license, attribution, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, 'Open CEDA', %s, %s, %s, FALSE)
                ON DUPLICATE KEY UPDATE
                    release_date = VALUES(release_date),
                    raw_factor_year = VALUES(raw_factor_year),
                    source_file_sha256 = VALUES(source_file_sha256),
                    source_license = VALUES(source_license),
                    attribution = VALUES(attribution)
                """,
                (
                    metadata.version_code,
                    metadata.release_date,
                    metadata.raw_factor_year,
                    metadata.reference_price_year,
                    metadata.currency_code,
                    metadata.price_basis,
                    metadata.source_file_sha256,
                    metadata.source_license,
                    metadata.attribution,
                ),
            )
            cursor.execute(
                """
                SELECT id FROM ceda_dataset_versions
                WHERE version_code = %s AND reference_price_year = %s
                  AND currency_code = %s AND price_basis = %s
                """,
                (
                    metadata.version_code,
                    metadata.reference_price_year,
                    metadata.currency_code,
                    metadata.price_basis,
                ),
            )
            dataset_id = int(cursor.fetchone()[0])

            cursor.executemany(
                """
                INSERT INTO ceda_countries (country_code, country_name)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE country_name = VALUES(country_name)
                """,
                countries,
            )
            for batch in _batches(sectors, batch_size):
                cursor.executemany(
                    """
                    INSERT INTO ceda_sectors
                        (sector_code, sector_name, naics_code, sector_description)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        sector_name = VALUES(sector_name),
                        naics_code = VALUES(naics_code),
                        sector_description = VALUES(sector_description)
                    """,
                    batch,
                )

            cursor.execute(
                "DELETE FROM ceda_emission_factors WHERE dataset_version_id = %s",
                (dataset_id,),
            )
            factor_values = [
                (
                    dataset_id,
                    row.country_code,
                    row.sector_code,
                    metadata.reference_price_year,
                    metadata.currency_code,
                    metadata.price_basis,
                    row.factor_value,
                    "kgCO2e/SGD",
                    row.raw_factor_value,
                    row.purchaser_conversion,
                    row.sector_price_index,
                    row.exchange_rate_lcu_per_usd,
                )
                for row in factors
            ]
            for batch in _batches(factor_values, batch_size):
                cursor.executemany(
                    """
                    INSERT INTO ceda_emission_factors
                        (dataset_version_id, country_code, sector_code,
                         reference_price_year, currency_code, price_basis,
                         factor_value, factor_unit, raw_factor_value,
                         purchaser_conversion, sector_price_index,
                         exchange_rate_lcu_per_usd)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    batch,
                )

            cursor.execute("UPDATE ceda_dataset_versions SET is_active = FALSE")
            cursor.execute(
                "UPDATE ceda_dataset_versions SET is_active = TRUE WHERE id = %s",
                (dataset_id,),
            )
            conn.commit()
            return len(factors)
        except mysql.connector.Error:
            conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import country-specific 2025 SGD purchaser-price factors from Open CEDA."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--reference-price-year", type=int, default=2025)
    parser.add_argument("--batch-size", type=int, default=1000)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and extract factors without writing to the database.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    metadata, factors = extract_open_ceda(
        args.workbook,
        reference_price_year=args.reference_price_year,
    )
    countries = {row.country_code for row in factors}
    sectors = {row.sector_code for row in factors}
    example = next(
        (
            row.factor_value
            for row in factors
            if row.country_code == "CHN" and row.sector_code == "331313"
        ),
        None,
    )
    print(
        f"Validated {metadata.version_code}: {len(countries)} countries, "
        f"{len(sectors)} sectors, {len(factors)} factors."
    )
    if example is not None:
        print(f"Check CHN/331313 = {example:.6f} kgCO2e/SGD.")
    if args.dry_run:
        return

    imported = import_open_ceda(metadata, factors, args.batch_size)
    print(f"Imported {imported} factors and activated {metadata.version_code}.")


if __name__ == "__main__":
    main()
