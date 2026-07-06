from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python package 'openpyxl'. Run this script with the project virtual environment:\n"
        r"  .\.venv\Scripts\python.exe scripts\import_machine_profiles.py --workbook DB\Machining_Equipment_Power_Emissions_SG.xlsx --dry-run"
    ) from exc

ROOT_DIR = Path(__file__).resolve().parents[1]
API_DIR = ROOT_DIR / "api"
if str(API_DIR) not in sys.path:
    sys.path.insert(0, str(API_DIR))


GRID_UPSERT_SQL = """
INSERT INTO method2_grid_electricity_factors
    (country_code, region_name, year, kgco2e_per_kwh, data_source)
VALUES
    (%s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
    kgco2e_per_kwh = VALUES(kgco2e_per_kwh),
    data_source = VALUES(data_source)
"""


MACHINE_UPSERT_SQL = """
INSERT INTO method2_machine_profiles
    (
        machine_key,
        machine_name,
        duty_level,
        peak_power_kw,
        avg_operating_load_kw,
        voltage_v,
        frequency_hz,
        full_load_current_a,
        country_code,
        data_source
    )
VALUES
    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
    machine_name = VALUES(machine_name),
    peak_power_kw = VALUES(peak_power_kw),
    avg_operating_load_kw = VALUES(avg_operating_load_kw),
    voltage_v = VALUES(voltage_v),
    frequency_hz = VALUES(frequency_hz),
    full_load_current_a = VALUES(full_load_current_a),
    data_source = VALUES(data_source)
"""


@dataclass(frozen=True)
class MachineProfile:
    machine_key: str
    machine_name: str
    duty_level: str
    peak_power_kw: float
    avg_operating_load_kw: float
    voltage_v: float | None
    frequency_hz: float | None
    full_load_current_a: float | None


def slugify(value: str) -> str:
    cleaned = re.sub(r"^\s*\d+\.\s*", "", value.strip())
    cleaned = re.sub(r"[^a-z0-9]+", "_", cleaned.casefold()).strip("_")
    if not cleaned:
        raise ValueError(f"Cannot create machine key from {value!r}")
    return cleaned


def clean_machine_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^\s*\d+\.\s*", "", text)
    if not text:
        raise ValueError("Machine name is empty")
    return text


def optional_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def required_float(value: Any, field_name: str) -> float:
    parsed = optional_float(value)
    if parsed is None:
        raise ValueError(f"Missing numeric value for {field_name}")
    return parsed


def load_grid_factor(workbook_path: Path) -> float:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Summary & Parameters"]
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        label = str(row[0] or "")
        if "Singapore Grid Emission Factor" in label:
            return required_float(row[1], "Singapore grid emission factor")
    raise ValueError("Could not find Singapore grid emission factor in Summary & Parameters")


def load_machine_profiles(workbook_path: Path, grid_factor: float, tolerance: float) -> tuple[list[MachineProfile], list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Equipment Specifications"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    profiles: list[MachineProfile] = []
    warnings: list[str] = []

    for index, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        machine_raw = row.get("Machine Type")
        duty_raw = row.get("Duty Level")
        if machine_raw is None or duty_raw is None:
            continue

        machine_text = str(machine_raw).strip()
        duty_level = str(duty_raw).strip()
        if not re.match(r"^\s*\d+\.", machine_text) or "Duty" not in duty_level:
            continue

        machine_name = clean_machine_name(machine_text)
        avg_kw = required_float(row.get("Avg Operating Load (kW)"), "Avg Operating Load (kW)")
        hourly_from_excel = optional_float(row.get("Hourly Scope 2 Emissions (kg CO2e/hr)"))
        calculated_hourly = avg_kw * grid_factor
        if hourly_from_excel is not None and abs(calculated_hourly - hourly_from_excel) > tolerance:
            warnings.append(
                f"Row {index}: {machine_name} / {duty_level} Excel hourly "
                f"{hourly_from_excel:.4f} differs from avg kW * grid factor {calculated_hourly:.4f}."
            )

        profiles.append(
            MachineProfile(
                machine_key=slugify(machine_name),
                machine_name=machine_name,
                duty_level=duty_level,
                peak_power_kw=required_float(row.get("Peak Wattage (kW)"), "Peak Wattage (kW)"),
                avg_operating_load_kw=avg_kw,
                voltage_v=optional_float(row.get("Voltage (V)")),
                frequency_hz=optional_float(row.get("Frequency (Hz)")),
                full_load_current_a=optional_float(row.get("Full Load Current (A)")),
            )
        )

    if not profiles:
        raise ValueError("No valid machine-duty rows found in Equipment Specifications")
    return profiles, warnings


def seed_database(
    profiles: list[MachineProfile],
    grid_factor: float,
    country_code: str,
    year: int,
    data_source: str,
) -> None:
    import mysql.connector
    from db import get_conn

    conn = get_conn()
    try:
        cursor = conn.cursor()
        try:
            cursor.execute(
                GRID_UPSERT_SQL,
                (country_code, "Singapore", year, grid_factor, data_source),
            )
            cursor.executemany(
                MACHINE_UPSERT_SQL,
                [
                    (
                        profile.machine_key,
                        profile.machine_name,
                        profile.duty_level,
                        profile.peak_power_kw,
                        profile.avg_operating_load_kw,
                        profile.voltage_v,
                        profile.frequency_hz,
                        profile.full_load_current_a,
                        country_code,
                        data_source,
                    )
                    for profile in profiles
                ],
            )
            conn.commit()
        except mysql.connector.Error:
            conn.rollback()
            raise
        finally:
            cursor.close()
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load Method 2 machine profiles from the company Excel workbook.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path.home() / "Downloads" / "Machining_Equipment_Power_Emissions_SG.xlsx",
        help="Company workbook containing Summary & Parameters and Equipment Specifications sheets.",
    )
    parser.add_argument("--country-code", default="SG", help="Country code for imported machine profiles.")
    parser.add_argument("--year", type=int, default=2026, help="Grid factor year to upsert.")
    parser.add_argument(
        "--data-source",
        default="Company workbook EMA 2025/2026",
        help="Source label stored with imported rows.",
    )
    parser.add_argument(
        "--tolerance",
        type=float,
        default=0.02,
        help="Allowed kg CO2e/hr difference between Excel hourly emissions and recalculated emissions.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate without writing to the database.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    grid_factor = load_grid_factor(args.workbook)
    profiles, warnings = load_machine_profiles(args.workbook, grid_factor, args.tolerance)

    print(f"Singapore grid factor: {grid_factor:.4f} kg CO2e/kWh")
    print(f"Prepared {len(profiles)} machine-duty profiles from {args.workbook}.")
    for warning in warnings:
        print(f"Warning: {warning}")

    if args.dry_run:
        return

    seed_database(profiles, grid_factor, args.country_code, args.year, args.data_source)
    print(f"Imported {len(profiles)} machine-duty profiles for {args.country_code}.")


if __name__ == "__main__":
    main()
